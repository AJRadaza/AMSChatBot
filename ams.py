# Python libraries that we need to import for our bot
from flask import Flask, request #Dependency
from pymessenger.bot import Bot #Dependency
from apscheduler.schedulers.background import BackgroundScheduler #Dependency
from datetime import date

import openpyxl #Dependency
import requests #Dependency
import atexit
import random
import time
import datetime
import json
import os
import calendar

# This is a helper function for accessing files
def get_dir(dir):
	curr_dir = os.path.dirname(__file__)
	return os.path.join(curr_dir, dir)

# Created this function as a helper that checks if a string represents an integer
def RepresentsInt(s):
	try:
		int(s)
		return True
	except ValueError:
		return False

# Helper function to convert list to string   
def listToString(s):
	str1 = ""
	for ele in s:
		str1 += ele
	
	return str1

# Helper function that removes duplicates in a list
def del_dup(x):
	return list(dict.fromkeys(x))

# Class for attendance management
class AttendanceData:
	def __init__(self, officer):
		self.week = 1
		self.absentees = {}
		self.tardies = {}
		self.data = {}
		self.officer = officer
		sched.add_job(func=self.set_week, trigger='cron', day_of_week='mon', hour=0, minute=0)

	def set_week(self, week=1):
		if week == 1:
			self.week += week
		else:
			self.week = week
		self.save()

	def mark_absent(self, name, subj_absent='all'):
		if name in self.tardies:
			self.tardies.pop(name)
		if subj_absent=='all':
			self.absentees[name] = [subj_absent]
		else:
			if len(subj_absent.split(',')) == 1:
				self.absentees[name] = [subj_absent]
			else:
				self.absentees[name] = [x.strip().lower() for x in subj_absent.split(',')]
		if not f'week{self.week}' in self.data:
			self.data[f'week{self.week}'] = {}
		if not calendar.day_name[datetime.datetime.today().weekday()] in self.data[f'week{self.week}']:
			self.data[f'week{self.week}'][calendar.day_name[datetime.datetime.today().weekday()]] = {}
		self.data[f'week{self.week}'][calendar.day_name[datetime.datetime.today().weekday()]]['absentees'] = self.absentees
		self.save()

	def mark_late(self, name, subj_came='all'):
		if name in self.absentees:
			self.absentees.pop(name)
		self.tardies[name] = subj_came
		if not f'week{self.week}' in self.data:
			self.data[f'week{self.week}'] = {}
		if not calendar.day_name[datetime.datetime.today().weekday()] in self.data[f'week{self.week}']:
			self.data[f'week{self.week}'][calendar.day_name[datetime.datetime.today().weekday()]] = {}
		self.data[f'week{self.week}'][calendar.day_name[datetime.datetime.today().weekday()]]['tardies'] = self.tardies
		self.save()

	def mark_present(self, name):
		if name in self.absentees:
			self.absentees.pop(name)
		if name in self.tardies:
			self.tardies.pop(name)
		self.save()

	def load(self):
		if os.path.exists(get_dir(f"data/attendance_data/officer-{self.officer.recipient_id}.json")):
			try:
				with open(get_dir(f"data/attendance_data/officer-{self.officer.recipient_id}.json")) as ff:
					self.data = json.load(ff)
					print(self.data)
					self.week = len(self.data)
					if self.week == 0:
						self.week = 1
					
					if not f'week{self.week}' in self.data:
						self.data[f'week{self.week}'] = {}
						print("Current week does not exist.")
					else:
						print("Current week exists.")
					if not calendar.day_name[datetime.datetime.today().weekday()] in self.data[f'week{self.week}']:
						self.data[f'week{self.week}'][calendar.day_name[datetime.datetime.today().weekday()]] = {}
						print("Current day does not exist.")
					else:
						print("Current day exists.")

					print(self.data)

					if 'tardies' in self.data[f'week{self.week}'][calendar.day_name[datetime.datetime.today().weekday()]]:
						self.tardies = self.data[f'week{self.week}'][calendar.day_name[datetime.datetime.today().weekday()]]['tardies']
					else:
						print('no tardies')

					if 'absentees' in self.data[f'week{self.week}'][calendar.day_name[datetime.datetime.today().weekday()]]:
						self.absentees = self.data[f'week{self.week}'][calendar.day_name[datetime.datetime.today().weekday()]]['absentees']
					else:
						print('no absentees')

					print("Loaded officer's attendance data.")
			except:
				print("The data in the attendance data is either corrupt or empty.")
		else:
			print("There is no attendance_data saved for the officer.")

	def save(self):
		if os.path.exists(get_dir(f"data/attendance_data/officer-{self.officer.recipient_id}.json")):
			with open(get_dir(f"data/attendance_data/officer-{self.officer.recipient_id}.json"), 'r+') as attendance_data:
				file_data = json.load(attendance_data)

				if f'week{self.week}' in file_data:
					print("Week is still current.")
					file_data.pop(f'week{self.week}')
					file_data[f'week{self.week}'] = self.data[f'week{self.week}']
				else:
					print("It's a new week!")
					file_data[f'week{self.week}'] = {}
					if not f'week{self.week}' in self.data:
						self.data[f'week{self.week}'] = {}
					if not calendar.day_name[datetime.datetime.today().weekday()] in self.data[f'week{self.week}']:
						self.data[f'week{self.week}'][calendar.day_name[datetime.datetime.today().weekday()]] = {}
					file_data[f'week{self.week}'] = self.data[f'week{self.week}']

				attendance_data.seek(0)
				json.dump(file_data, attendance_data, indent=4)
				attendance_data.truncate()
				print("AttendanceData was successfully loaded, modified and saved.")
		else:
			with open(get_dir(f"data/attendance_data/officer-{self.officer.recipient_id}.json"), 'w') as attendance_data:
				file_data = self.data
				json.dump(file_data, attendance_data, indent=4)
				print("Attendance data was successfully saved.")

# Created this class for object-oriented data management for all class officers
class Officer:
	def __init__(self, recipient_id):
		self.recipient_id = recipient_id
		self.mwf_subjects = list()
		self.tth_subjects = list()
		self.students_male = list()
		self.students_female = list()
		self.is_registered = False
		self.officer_data = {}
		self.attendance_data = AttendanceData(self)

	def load(self):
		try:
			with open(get_dir(f"data/officer_data/officer-{self.recipient_id}.json")) as f:
				self.officer_data = json.load(f)
				self.recipient_id = self.officer_data['id']
				print(self.recipient_id)
				print('loaded recipient_id')
				self.name = self.officer_data['name']
				print(self.name)
				print('loaded name.')
				self.grade_level = self.officer_data['grade_level']
				print(self.grade_level)
				print('loaded grade level')
				self.class_section = self.officer_data['class_section']
				print(self.class_section)
				print("loaded section.")
				self.class_adviser = self.officer_data['class_adviser']
				print(self.class_adviser)
				print("loaded adviser.")
				self.mwf_subjects = self.officer_data['mwf_subjects']
				print(self.mwf_subjects)
				print('loaded mwf')
				self.tth_subjects = self.officer_data['tth_subjects']
				print(self.tth_subjects)
				print("loaded tth")
				self.mwf_time = self.officer_data['mwf_time']
				print(self.mwf_time)
				print("loaded mwf time")
				self.tth_time = self.officer_data['tth_time']
				print(self.tth_time)
				print("loaded tth time")
				self.students_male = self.officer_data['students_male']
				print(self.students_male)
				print('loaded male students')
				self.students_female = self.officer_data['students_female']
				print(self.students_female)
				print('loaded female students')
				self.attendance_data.load()
				print(self.attendance_data.absentees)
				print(self.attendance_data.tardies)
				self.is_registered = True
		except IOError:
			print('File unaccessible.')

	def initial_reg(self, name):
		self.name = name

	def register_grade(self, grade_level):
		self.grade_level = grade_level

	def register_section(self, class_section):
		self.class_section = class_section

	def register_adviser(self, class_adviser):
		self.class_adviser = class_adviser

	def register_mwf(self, mwf_subjects):
		self.mwf_subjects = mwf_subjects

	def register_tth(self, tth_subjects):
		self.tth_subjects = tth_subjects

	def register_mwf_time(self, mwf_time):
		self.mwf_time = mwf_time

	def register_tth_time(self, tth_time):
		self.tth_time = tth_time

	def register_student_male(self, student):
		self.students_male.append(student)

	def register_student_female(self, student):
		self.students_female.append(student)

	def register_complete(self, client):
		self.officer_data['id'] = self.recipient_id
		print('saved recipient_id')

		self.officer_data['name'] = self.name
		print('saved name.')

		self.officer_data['grade_level'] = self.grade_level
		print('saved grade level')

		self.officer_data['class_section'] = self.class_section
		print("saved section.") 

		self.officer_data['class_adviser'] = self.class_adviser
		print("saved adviser.")

		self.officer_data['mwf_subjects'] = self.mwf_subjects
		print('saved mwf')

		self.officer_data['tth_subjects'] = self.tth_subjects
		print("saved tth")

		self.officer_data['mwf_time'] = self.mwf_time
		print("saved mwf time")

		self.officer_data['tth_time'] = self.tth_time
		print("saved tth time")

		self.students_male.sort(key=lambda x:x['last'])
		self.officer_data['students_male'] = self.students_male
		print('saved male students')

		self.students_female.sort(key=lambda x:x['last'])
		self.officer_data['students_female'] = self.students_female
		print('saved female students')

		print('Applying to scheduler...')
		time_mwf = self.mwf_time.split(':')
		hour_mwf = int(time_mwf[0])
		minute_mwf = int(time_mwf[1])
		print(time_mwf)
		print(hour_mwf)
		print(minute_mwf)
			
		sched.add_job(func=schedule_officer, trigger='cron', args=[self.recipient_id], day_of_week='mon', hour=hour_mwf, minute=minute_mwf)
		sched.add_job(func=schedule_officer, trigger='cron', args=[self.recipient_id], day_of_week='wed', hour=hour_mwf, minute=minute_mwf)
		sched.add_job(func=schedule_officer, trigger='cron', args=[self.recipient_id], day_of_week='fri', hour=hour_mwf, minute=minute_mwf)

		# TTH
		time_tth = self.tth_time.split(':')
		hour_tth = int(time_tth[0])
		minute_tth = int(time_tth[1])
		print(time_tth)
		print(hour_tth)
		print(minute_tth)

		sched.add_job(func=schedule_officer, trigger='cron', args=[self.recipient_id], day_of_week='tue', hour=hour_tth, minute=minute_tth)
		sched.add_job(func=schedule_officer, trigger='cron', args=[self.recipient_id], day_of_week='thu', hour=hour_tth, minute=minute_tth)

		sched.add_job(func=self.save_to_xlsx, trigger='cron', day_of_week='mon-fri', hour=18, minute=0)
		sched.add_job(func=self.send_data_to_user, trigger='cron', args=[client], day_of_week='fri', hour=18, minute=0)

		print("Applied to scheduler.")

		print("Saving attendance_data.")
		self.attendance_data.save()

		self.is_registered = True
		print("data ready for export.")

	def save_to_file(self):
		print("Trying to save...")

		if self.is_registered:
			print(self.officer_data)
			print("Saving file...")
			with open(get_dir(f"data/officer_data/officer-{self.recipient_id}.json"), 'w') as f:
				json.dump(self.officer_data, f, indent=4)
				print("Json file saved.")

		if os.path.exists(get_dir("data/register.json")):
			try:
				with open(get_dir("data/register.json"), 'r+') as register_data:
					file_data = json.load(register_data)	
					file_data[f"{self.recipient_id}"] = {
						'mwf_time': self.mwf_time,
						'tth_time': self.tth_time
					}
					register_data.seek(0)
					json.dump(file_data, register_data, indent=4)
					register_data.truncate()
					print("Register data was successfully loaded, modified and saved.")
			except IOError:
				print("File not accessible")
		else:	
			with open(get_dir("data/register.json"), 'w') as register_data:
				file_data = {}
				file_data[f"{self.recipient_id}"] = {
					'mwf_time': self.mwf_time,
					'tth_time': self.tth_time
				}
				json.dump(file_data, register_data, indent=4)
				print("Register data was successfully saved.")

		with open(get_dir(f"data/attendance_data/officer-{self.recipient_id}.json"), 'w') as ff:
			json.dump(self.attendance_data.data, ff, indent=4)
			print("Saved empty attendance data.")

	def save_to_xlsx(self):	
		print("Saving to xlsx format...")

		officer_data = self.officer_data
		attendance_data = self.attendance_data
		attendance_data_json = self.attendance_data.data

		xfile = openpyxl.load_workbook(get_dir("data/attendance.xlsx"))
		sheet = xfile['Sheet1']

		# Here is where you assign the data gathered from the officer to the excel file.

		height_page_1 = float(0)
		for hw in range(57):
			height_page_1 += float(sheet.row_dimensions[hw+1].height)

		print(f"Height of page 1: {height_page_1}")
		
		# Class Grade and Section
		sheet['A3'] = f"Grade {officer_data['grade_level']} - {officer_data['class_section']}"

		# Class male students
		male_students = officer_data['students_male']

		for male in range(len(male_students)):
			sheet[f"B{8+male}"] = male_students[male].get('last')
			sheet[f"D{8+male}"] = male_students[male].get('first')
			sheet[f"E{8+male}"] = male_students[male].get('middle')

		if len(male_students) < 50:
			sheet.delete_rows(8+len(male_students), 50-len(male_students))

		# Class female students
		female_students = officer_data['students_female']

		female_row = 8+len(male_students)

		if female_row < 58:
			sheet.merge_cells(f"A{female_row}:E{female_row}")
			sheet.row_dimensions[female_row].height = 70

		for female in range(len(female_students)):
			sheet[f"A{female_row+1+female}"] = str(len(male_students) + female + 1)
			sheet[f"B{female_row+1+female}"] = female_students[female].get('last')
			sheet[f"D{female_row+1+female}"] = female_students[female].get('first')
			sheet[f"E{female_row+1+female}"] = female_students[female].get('middle')

		if len(female_students) < 50:
			sheet.delete_rows(8+len(male_students)+1+len(female_students), 50-len(female_students))

		# Class subjects
		mwf_subjects = officer_data['mwf_subjects']
		tth_subjects = officer_data['tth_subjects']

		for mwf in range(len(mwf_subjects)):
			sheet.cell(row = 7, column = 6+mwf).value = mwf_subjects[mwf]
			sheet.cell(row = 7, column = 26+mwf).value = mwf_subjects[mwf]
			sheet.cell(row = 7, column = 46+mwf).value = mwf_subjects[mwf]
			sheet.cell(row = female_row, column = 6+mwf).value = mwf_subjects[mwf]
			sheet.cell(row = female_row, column = 26+mwf).value = mwf_subjects[mwf]
			sheet.cell(row = female_row, column = 46+mwf).value = mwf_subjects[mwf]

		for tth in range(len(tth_subjects)):
			sheet.cell(row = 7, column = 16+tth).value = tth_subjects[tth]
			sheet.cell(row = 7, column = 36+tth).value = tth_subjects[tth]
			sheet.cell(row = female_row, column = 16+tth).value = tth_subjects[tth]
			sheet.cell(row = female_row, column = 36+tth).value = tth_subjects[tth]

		# Here is where you input the gathered attendance data.
		if f'week{attendance_data.week}' in attendance_data_json:
			curr_week = attendance_data_json[f'week{attendance_data.week}']
			# MWF
			if 'Monday' in curr_week:
				curr_day = curr_week['Monday']
				for subs in range(len(mwf_subjects)):
					for mmm in range(len(male_students)):
						sheet.cell(row = 8+mmm, column = 6+subs).value = '/'
					for fff in range(len(female_students)):
						sheet.cell(row = female_row+1+fff, column = 6+subs).value = '/'
				if 'absentees' in curr_day:
					for ttt,kkk in curr_day['absentees'].items():
						if len(kkk) == 1:
							if kkk[0] == 'all':
								for mmm in range(len(male_students)):
									if ttt == male_students[mmm].get('last').lower():
										for columnnn in range(len(mwf_subjects)):
											sheet.cell(row = 8+mmm, column = 6 + columnnn).value = 'X'
								for fff in range(len(female_students)):
									if ttt == female_students[fff].get('last').lower():
										for columnnn in range(len(mwf_subjects)):
											sheet.cell(row = female_row+1+fff, column = 6 + columnnn).value = 'X'
							else:
								for mmm in range(len(male_students)):
									if ttt == male_students[mmm].get('last').lower():
										for columnnn in range(len(mwf_subjects)):
											if mwf_subjects[columnnn].lower() in kkk:
												sheet.cell(row = 8+mmm, column = 6 + columnnn).value = 'X'
								for fff in range(len(female_students)):
									if ttt == female_students[fff].get('last').lower():
										for columnnn in range(len(mwf_subjects)):
											if mwf_subjects[columnnn].lower() in kkk:
												sheet.cell(row = female_row+1+fff, column = 6 + columnnn).value = 'X'
				if 'tardies' in curr_day:
					for ttt,kkk in curr_day['tardies'].items():
						if kkk == 'all':
							for mmm in range(len(male_students)):
								if ttt == male_students[mmm].get('last').lower():
									sheet.cell(row = 8+mmm, column = 6).value = 'L'
							for fff in range(len(female_students)):
								if ttt == female_students[fff].get('last').lower():
									sheet.cell(row = female_row+1+fff, column = 6).value = 'L'
						else:
							print('Late in more than 1 subject.')
							subj_loc = 0
							for subsss in range(len(mwf_subjects)):
								print(subsss)
								if mwf_subjects[subsss].lower() == kkk:
									subj_loc = subsss + 1
									print("Set subj_loc.")
							print(subj_loc)
							for mmm in range(len(male_students)):
								if ttt == male_students[mmm].get('last').lower():
									for columnnn in range(subj_loc):
										sheet.cell(row = 8+mmm, column=6+columnnn).value = 'X'
										if columnnn + 1 == subj_loc:
											sheet.cell(row = 8+mmm, column = 6+columnnn).value = 'L'
							for fff in range(len(female_students)):
								if ttt == female_students[fff].get('last').lower():
									for columnnn in range(subj_loc):
										sheet.cell(row = 8+mmm, column=6+columnnn).value = 'X'
										if columnnn + 1 == subj_loc:
											sheet.cell(row = female_row+1+fff, column = 6+columnnn).value = 'L'
			if 'Wednesday' in curr_week:
				curr_day = curr_week['Wednesday']
				for subs in range(len(mwf_subjects)):
					for mmm in range(len(male_students)):
						sheet.cell(row = 8+mmm, column = 26+subs).value = '/'
					for fff in range(len(female_students)):
						sheet.cell(row = female_row+1+fff, column = 26+subs).value = '/'
				if 'absentees' in curr_day:
					for ttt,kkk in curr_day['absentees'].items():
						if len(kkk) == 1:
							if kkk[0] == 'all':
								for mmm in range(len(male_students)):
									if ttt == male_students[mmm].get('last').lower():
										for columnnn in range(len(mwf_subjects)):
											sheet.cell(row = 8+mmm, column = 26 + columnnn).value = 'X'
								for fff in range(len(female_students)):
									if ttt == female_students[fff].get('last').lower():
										for columnnn in range(len(mwf_subjects)):
											sheet.cell(row = female_row+1+fff, column = 26 + columnnn).value = 'X'
							else:
								for mmm in range(len(male_students)):
									if ttt == male_students[mmm].get('last').lower():
										for columnnn in range(len(mwf_subjects)):
											if mwf_subjects[columnnn].lower() in kkk:
												sheet.cell(row = 8+mmm, column = 26 + columnnn).value = 'X'
								for fff in range(len(female_students)):
									if ttt == female_students[fff].get('last').lower():
										for columnnn in range(len(mwf_subjects)):
											if mwf_subjects[columnnn].lower() in kkk:
												sheet.cell(row = female_row+1+fff, column = 26 + columnnn).value = 'X'
				if 'tardies' in curr_day:
					for ttt,kkk in curr_day['tardies'].items():
						if kkk == 'all':
							for mmm in range(len(male_students)):
								if ttt == male_students[mmm].get('last').lower():
									sheet.cell(row = 8+mmm, column = 26).value = 'L'
							for fff in range(len(female_students)):
								if ttt == female_students[fff].get('last').lower():
									sheet.cell(row = female_row+1+fff, column = 26).value = 'L'
						else:
							print('Late in more than 1 subject.')
							subj_loc = 0
							for subsss in range(len(mwf_subjects)):
								print(subsss)
								if mwf_subjects[subsss].lower() == kkk:
									subj_loc = subsss + 1
									print("Set subj_loc.")
							print(subj_loc)
							for mmm in range(len(male_students)):
								if ttt == male_students[mmm].get('last').lower():
									for columnnn in range(subj_loc):
										sheet.cell(row = 8+mmm, column=26+columnnn).value = 'X'
										if columnnn + 1 == subj_loc:
											sheet.cell(row = 8+mmm, column = 26+columnnn).value = 'L'
							for fff in range(len(female_students)):
								if ttt == female_students[fff].get('last').lower():
									for columnnn in range(subj_loc):
										sheet.cell(row = 8+mmm, column=26+columnnn).value = 'X'
										if columnnn + 1 == subj_loc:
											sheet.cell(row = female_row+1+fff, column = 26+columnnn).value = 'L'
			if 'Friday' in curr_week:
				curr_day = curr_week['Friday']
				for subs in range(len(mwf_subjects)):
					for mmm in range(len(male_students)):
						sheet.cell(row = 8+mmm, column = 46+subs).value = '/'
					for fff in range(len(female_students)):
						sheet.cell(row = female_row+1+fff, column = 46+subs).value = '/'
				if 'absentees' in curr_day:
					for ttt,kkk in curr_day['absentees'].items():
						if len(kkk) == 1:
							if kkk[0] == 'all':
								for mmm in range(len(male_students)):
									if ttt == male_students[mmm].get('last').lower():
										for columnnn in range(len(mwf_subjects)):
											sheet.cell(row = 8+mmm, column = 46 + columnnn).value = 'X'
								for fff in range(len(female_students)):
									if ttt == female_students[fff].get('last').lower():
										for columnnn in range(len(mwf_subjects)):
											sheet.cell(row = female_row+1+fff, column = 46 + columnnn).value = 'X'
							else:
								for mmm in range(len(male_students)):
									if ttt == male_students[mmm].get('last').lower():
										for columnnn in range(len(mwf_subjects)):
											if mwf_subjects[columnnn].lower() in kkk:
												sheet.cell(row = 8+mmm, column = 46 + columnnn).value = 'X'
								for fff in range(len(female_students)):
									if ttt == female_students[fff].get('last').lower():
										for columnnn in range(len(mwf_subjects)):
											if mwf_subjects[columnnn].lower() in kkk:
												sheet.cell(row = female_row+1+fff, column = 46 + columnnn).value = 'X'
				if 'tardies' in curr_day:
					for ttt,kkk in curr_day['tardies'].items():
						if kkk == 'all':
							for mmm in range(len(male_students)):
								if ttt == male_students[mmm].get('last').lower():
									sheet.cell(row = 8+mmm, column = 46).value = 'L'
							for fff in range(len(female_students)):
								if ttt == female_students[fff].get('last').lower():
									sheet.cell(row = female_row+1+fff, column = 46).value = 'L'
						else:
							print('Late in more than 1 subject.')
							subj_loc = 0
							for subsss in range(len(mwf_subjects)):
								print(subsss)
								if mwf_subjects[subsss].lower() == kkk:
									subj_loc = subsss + 1
									print("Set subj_loc.")
							print(subj_loc)
							for mmm in range(len(male_students)):
								if ttt == male_students[mmm].get('last').lower():
									for columnnn in range(subj_loc):
										sheet.cell(row = 8+mmm, column=46+columnnn).value = 'X'
										if columnnn + 1 == subj_loc:
											sheet.cell(row = 8+mmm, column = 46+columnnn).value = 'L'
							for fff in range(len(female_students)):
								if ttt == female_students[fff].get('last').lower():
									for columnnn in range(subj_loc):
										sheet.cell(row = 8+mmm, column=46+columnnn).value = 'X'
										if columnnn + 1 == subj_loc:
											sheet.cell(row = female_row+1+fff, column = 46+columnnn).value = 'L'

			# TTH
			if 'Tuesday' in curr_week:
				curr_day = curr_week['Tuesday']
				for subs in range(len(tth_subjects)):
					for mmm in range(len(male_students)):
						sheet.cell(row = 8+mmm, column = 16+subs).value = '/'
					for fff in range(len(female_students)):
						sheet.cell(row = female_row+1+fff, column = 16+subs).value = '/'
				if 'absentees' in curr_day:
					for ttt,kkk in curr_day['absentees'].items():
						if len(kkk) == 1:
							if kkk[0] == 'all':
								for mmm in range(len(male_students)):
									if ttt == male_students[mmm].get('last').lower():
										for columnnn in range(len(tth_subjects)):
											sheet.cell(row = 8+mmm, column = 16 + columnnn).value = 'X'
								for fff in range(len(female_students)):
									if ttt == female_students[fff].get('last').lower():
										for columnnn in range(len(tth_subjects)):
											sheet.cell(row = female_row+1+fff, column = 16 + columnnn).value = 'X'
							else:
								for mmm in range(len(male_students)):
									if ttt == male_students[mmm].get('last').lower():
										for columnnn in range(len(tth_subjects)):
											if tth_subjects[columnnn].lower() in kkk:
												sheet.cell(row = 8+mmm, column = 16 + columnnn).value = 'X'
								for fff in range(len(female_students)):
									if ttt == female_students[fff].get('last').lower():
										for columnnn in range(len(tth_subjects)):
											if tth_subjects[columnnn].lower() in kkk:
												sheet.cell(row = female_row+1+fff, column = 16 + columnnn).value = 'X'
				if 'tardies' in curr_day:
					for ttt,kkk in curr_day['tardies'].items():
						if kkk == 'all':
							for mmm in range(len(male_students)):
								if ttt == male_students[mmm].get('last').lower():
									sheet.cell(row = 8+mmm, column = 16).value = 'L'
							for fff in range(len(female_students)):
								if ttt == female_students[fff].get('last').lower():
									sheet.cell(row = female_row+1+fff, column = 16).value = 'L'
						else:
							print('Late in more than 1 subject.')
							subj_loc = 0
							for subsss in range(len(tth_subjects)):
								print(subsss)
								if tth_subjects[subsss].lower() == kkk:
									subj_loc = subsss + 1
									print("Set subj_loc.")
							print(subj_loc)
							for mmm in range(len(male_students)):
								if ttt == male_students[mmm].get('last').lower():
									for columnnn in range(subj_loc):
										sheet.cell(row = 8+mmm, column=16+columnnn).value = 'X'
										if columnnn + 1 == subj_loc:
											sheet.cell(row = 8+mmm, column = 16+columnnn).value = 'L'
							for fff in range(len(female_students)):
								if ttt == female_students[fff].get('last').lower():
									for columnnn in range(subj_loc):
										sheet.cell(row = 8+mmm, column=16+columnnn).value = 'X'
										if columnnn + 1 == subj_loc:
											sheet.cell(row = female_row+1+fff, column = 16+columnnn).value = 'L'
			if 'Thursday' in curr_week:
				curr_day = curr_week['Thursday']
				for subs in range(len(tth_subjects)):
					for mmm in range(len(male_students)):
						sheet.cell(row = 8+mmm, column = 36+subs).value = '/'
					for fff in range(len(female_students)):
						sheet.cell(row = female_row+1+fff, column = 36+subs).value = '/'
				if 'absentees' in curr_day:
					for ttt,kkk in curr_day['absentees'].items():
						if len(kkk) == 1:
							if kkk[0] == 'all':
								for mmm in range(len(male_students)):
									if ttt == male_students[mmm].get('last').lower():
										for columnnn in range(len(tth_subjects)):
											sheet.cell(row = 8+mmm, column = 36 + columnnn).value = 'X'
								for fff in range(len(female_students)):
									if ttt == female_students[fff].get('last').lower():
										for columnnn in range(len(tth_subjects)):
											sheet.cell(row = female_row+1+fff, column = 36 + columnnn).value = 'X'
							else:
								for mmm in range(len(male_students)):
									if ttt == male_students[mmm].get('last').lower():
										for columnnn in range(len(tth_subjects)):
											if tth_subjects[columnnn].lower() in kkk:
												sheet.cell(row = 8+mmm, column = 36 + columnnn).value = 'X'
								for fff in range(len(female_students)):
									if ttt == female_students[fff].get('last').lower():
										for columnnn in range(len(tth_subjects)):
											if tth_subjects[columnnn].lower() in kkk:
												sheet.cell(row = female_row+1+fff, column = 36 + columnnn).value = 'X'
				if 'tardies' in curr_day:
					for ttt,kkk in curr_day['tardies'].items():
						if kkk == 'all':
							for mmm in range(len(male_students)):
								if ttt == male_students[mmm].get('last').lower():
									sheet.cell(row = 8+mmm, column = 36).value = 'L'
							for fff in range(len(female_students)):
								if ttt == female_students[fff].get('last').lower():
									sheet.cell(row = female_row+1+fff, column = 36).value = 'L'
						else:
							print('Late in more than 1 subject.')
							subj_loc = 0
							for subsss in range(len(tth_subjects)):
								print(subsss)
								if tth_subjects[subsss].lower() == kkk:
									subj_loc = subsss + 1
									print("Set subj_loc.")
							print(subj_loc)
							for mmm in range(len(male_students)):
								if ttt == male_students[mmm].get('last').lower():
									for columnnn in range(subj_loc):
										sheet.cell(row = 8+mmm, column=36+columnnn).value = 'X'
										if columnnn + 1 == subj_loc:
											sheet.cell(row = 8+mmm, column = 36+columnnn).value = 'L'
							for fff in range(len(female_students)):
								if ttt == female_students[fff].get('last').lower():
									for columnnn in range(subj_loc):
										sheet.cell(row = 8+mmm, column=36+columnnn).value = 'X'
										if columnnn + 1 == subj_loc:
											sheet.cell(row = female_row+1+fff, column = 36+columnnn).value = 'L'


		# Merge class adviser
		sheet.merge_cells(f"A{female_row+len(female_students)+1}:BC{female_row+len(female_students)+1}")
		sheet.merge_cells(f"A{female_row+len(female_students)+2}:BC{female_row+len(female_students)+2}")

		sheet[f"A{female_row+len(female_students)+1}"] = officer_data['class_adviser']

		sheet.row_dimensions[female_row+len(female_students)+1].height = 25
		sheet.row_dimensions[female_row+len(female_students)+2].height = 22

		# Setting page area
		height_page_1_new = float(0)
		for hwn in range(female_row+len(female_students)+2):
			height_page_1_new += sheet.row_dimensions[hwn+1].height

		print(f"Height of page 1 new: {height_page_1_new}")
		if height_page_1 > height_page_1_new:
			sheet.print_area = f'A1:BC{female_row+len(female_students)+2}'

		# End
		xfile.save(get_dir("data/officer_data/export")+f"/officer-{officer_data['id'].translate({ord(i):None for i in ' '})}-week{self.attendance_data.week}.xlsx")
		print("Saved to xlsx.")

	def send_data_to_user(self, client, week=None):
		self.save_to_xlsx()
		print("Sending Xlsx to user...")
		if week:
			client.send_file(recipient_id=self.recipient_id, file_path=get_dir("data/officer_data/export")+f"/officer-{self.officer_data['id'].translate({ord(i):None for i in ' '})}-week{week}.xlsx")
		else:
			week = self.attendance_data.week
			client.send_file(recipient_id=self.recipient_id, file_path=get_dir("data/officer_data/export")+f"/officer-{self.officer_data['id'].translate({ord(i):None for i in ' '})}-week{week}.xlsx")
		print('Sent.')

# Also created this class for conversation handling system (Like taking inputs and giving outputs to and from a console)
class Conversation:
	def __init__(self, client, recipient_id, officer=None):
		self.client = client
		self.locale_strings = client.get_locale_strings('en')
		self.conv_reply = ""
		self.is_typing = False
		if officer == None:
			self.officer = Officer(recipient_id)
			self.recipient_id = recipient_id
			self.conv_state = 0
		else:
			self.officer = officer
			self.recipient_id = officer.recipient_id
			self.conv_state = 10

	def set_locale_strings(self, lang):
		if lang == 'en':
			self.locale_strings = self.client.get_locale_strings(lang)
		elif lang == 'bs':
			self.locale_strings = self.client.get_locale_strings(lang)
		else:
			print(f"\"{lang}\" is not a locale.")

	def convo(self):
		conv_reply = self.conv_reply
		conv_state = self.conv_state
		locale_strings = self.locale_strings

		print(f"Conversation reply: {conv_reply}")
		print(f"Conversation state: {conv_state}")

		if not self.officer.is_registered:
			if conv_state == 0:
				user_profile = get_profile(self.recipient_id)
				self.officer.initial_reg(f"{user_profile.get('first_name')} {user_profile.get('last_name')}")
				self.client_message(locale_strings['languageSelect'])
				self.is_typing = False
			elif conv_state == 1:
				officername = self.officer.name
				if conv_reply.lower() == "english":
					self.is_typing = True
					self.set_locale_strings('en')
					locale_strings = self.locale_strings
					self.client_message(locale_strings['langSelected'])
					# asyncio.sleep(1.5)
					self.client_message(locale_strings['greetingsUser'] % officername)
					# asyncio.sleep(2.5)
					self.client_message(locale_strings['fromNowOn'])
					# asyncio.sleep(2.5)
					self.client_message(locale_strings['justMakeSure'])
					# asyncio.sleep(2.5)
					self.client_message(locale_strings['beforeIHelp'])
					# asyncio.sleep(2.5)
					self.client_message(locale_strings['inWhatGrade'])
					self.is_typing = False
				elif conv_reply.lower() == "bisaya":
					self.is_typing = True
					self.set_locale_strings('bs')
					locale_strings = self.locale_strings
					self.client_message(locale_strings['langSelected'])
					# asyncio.sleep(1.5)
					self.client_message(locale_strings['greetingsUser'])
					# asyncio.sleep(2.5)
					self.client_message(locale_strings['fromNowOn'])
					# asyncio.sleep(2.5)
					self.client_message(locale_strings['justMakeSure'])
					# asyncio.sleep(2.5)
					self.client_message(locale_strings['beforeIHelp'])
					# asyncio.sleep(2.5)
					self.client_message(locale_strings['inWhatGrade'])
					self.is_typing = False
				else:
					self.client_message(locale_strings['enOrBs'])
					self.back_state()
			elif conv_state == 2:
				if RepresentsInt(conv_reply):
					if int(conv_reply) >= 7 and int(conv_reply) <= 12:
						self.officer.register_grade(conv_reply)
						print(self.officer.grade_level)
						self.client_message(locale_strings['whatSection'])
					else:
						self.client_message(locale_strings['choose712'])
						self.back_state()
				else:
					self.client_message(locale_strings['notGradeLevel'])
					self.back_state()
			elif conv_state == 3:
				self.officer.register_section(conv_reply)
				print(self.officer.class_section)
				self.client_message(locale_strings['classAdviser'])
			elif conv_state == 4:
				self.officer.register_adviser(conv_reply)
				print(self.officer.class_adviser)
				self.client_message(locale_strings['subjectMWF'])
			elif conv_state == 5:
				subject_mwf = conv_reply.split(',')
				for x in range(len(subject_mwf)):
					y = subject_mwf[x].strip()
					subject_mwf[x] = y
				self.officer.register_mwf(subject_mwf)
				print(self.officer.mwf_subjects)
				self.client_message(locale_strings['subjectTTH'])
			elif conv_state == 6:
				subject_tth = conv_reply.split(',')
				for x in range(len(subject_tth)):
					y = subject_tth[x].strip()
					subject_tth[x] = y
				self.officer.register_tth(subject_tth)
				print(self.officer.tth_subjects)
				self.client_message(locale_strings['firstSubjectTimeMWF'])
			elif conv_state == 7:
				time = conv_reply.split(":")
				hour = time[0]

				if len(time) == 2 and RepresentsInt(hour) and len(hour) in [1,2] and 1 <= int(hour) <= 12:
					if len(time[1]) == 4:					
						minute = time[1][0:2]
						ampm = time[1][2:]
						ampm = ampm.strip()
						if RepresentsInt(minute) and len(minute) == 2 and 0 <= int(minute) <= 60 and ampm in ["am", "pm"]:
							hour = int(hour)
							minute = int(minute)
							if ampm == "am":
								if hour == 12:
									hour = 0
							elif ampm == "pm":
								if hour > 12:
									hour = hour + 12
								elif hour == 12:
									hour = 12
							print(f"{hour}:{minute} {ampm}")
							self.officer.register_mwf_time(f"{hour}:{minute}")

							self.client_message(locale_strings['firstSubjectTimeTTH'])
						else:
							self.back_state()
							self.client_message(locale_strings['wrongTimeFormat'])				
							print("Time format wrong.")
					else:
						self.back_state()
						self.client_message(locale_strings['wrongTimeFormat'])
						print("Time format wrong.")
				else:
					self.back_state()
					self.client_message(locale_strings['wrongTimeFormat'])
					print("Time format wrong.")
			elif conv_state == 8:
				time = conv_reply.split(":")
				hour = time[0]

				if len(time) == 2 and RepresentsInt(hour) and len(hour) in [1,2] and 1 <= int(hour) <= 12:
					if len(time[1]) == 4:					
						minute = time[1][0:2]
						ampm = time[1][2:]
						ampm = ampm.strip()
						if RepresentsInt(minute) and len(minute) == 2 and 0 <= int(minute) <= 60 and ampm in ["am", "pm"]:
							hour = int(hour)
							minute = int(minute)
							print(f"{hour}:{minute} {ampm}")
							if ampm == "am":
								if hour == 12:
									hour = 0
							elif ampm == "pm":
								if hour > 12:
									hour += 12
							print(f"{hour}:{minute} {ampm}")
							self.officer.register_tth_time(f"{hour}:{minute}")
							
							self.client_message(locale_strings['studentNames'])
							# asyncio.sleep(1.5)
							self.client_message(locale_strings['studentReg'])
							# asyncio.sleep(1.5)
							self.client_message(locale_strings['studentRegMale'])
						else:
							self.back_state()
							self.client_message(locale_strings['wrongTimeFormat'])				
							print("Time format wrong.")
					else:
						self.back_state()
						self.client_message(locale_strings['wrongTimeFormat'])
						print("Time format wrong.")
				else:
					self.back_state()
					self.client_message(locale_strings['wrongTimeFormat'])
					print("Time format wrong.")
			elif conv_state == 9:
				if conv_reply.lower() == "done":
					"""print("Sending officer data...")
					self.officer.register_complete()
					self.officer.send_data_to_user(self.client, self.recipient_id)
					print("Sent.")"""
					self.client_message(locale_strings['maleReg'])
					# asyncio.sleep(1.5)
					self.client_message(locale_strings['studentRegFemale'])
				else:
					self.back_state()
					studentname = conv_reply.split(',')
					student_name = {}

					print(f"Length: {len(studentname)}")
					if len(studentname) > 1 and len(studentname) < 4:
						for x in range(len(studentname)):
							y = studentname[x].strip()
							studentname[x] = y
							if x == 0:
								student_name['last'] = studentname[x]
							elif x == 1:
								student_name['first'] = studentname[x]
							elif x == 2:
								student_name['middle'] = studentname[x]
						if len(studentname) == 2:
							student_name['middle'] = "-"
						print(student_name)
						self.officer.register_student_male(student_name)
						self.client_message("Registered.")
						if len(self.officer.students_male) == 1:
							# asyncio.sleep(10)
							self.client_message(locale_strings['ifDoneRegMale'])
					elif len(studentname) > 3:
						self.client_message(locale_strings['tooManyArgs'])
					elif len(studentname) == 1:
						self.client_message(locale_strings['lastNameFirstName'])
			elif conv_state == 10:
				if conv_reply.lower() == "done":
					self.officer.register_complete(self)
					self.officer.save_to_file()
					self.client_message(locale_strings['femaleReg'])
					self.client_message(locale_strings['thankYouReg'])
					self.client_message(locale_strings['emptyXlsx'])
					self.officer.send_data_to_user(self.client)
					self.client_message(locale_strings['canReset'])
					self.client_message("Type anything to proceed.")

					# asyncio.sleep(1.5)
				else:
					self.back_state()
					studentname = conv_reply.split(',')
					student_name = {}

					print(f"Length: {len(studentname)}")
					if len(studentname) > 1 and len(studentname) < 4:
						for x in range(len(studentname)):
							y = studentname[x].strip()
							studentname[x] = y
							if x == 0:
								student_name['last'] = studentname[x]
							elif x == 1:
								student_name['first'] = studentname[x]
							elif x == 2:
								student_name['middle'] = studentname[x]
						if len(studentname) == 2:
							student_name['middle'] = "-"
						print(student_name)
						self.officer.register_student_female(student_name)
						self.client_message("Registered.")
						if len(self.officer.students_female) == 1:
							#asyncio.sleep(10)
							self.client_message(locale_strings['ifDoneRegFemale'])
					elif len(studentname) > 3:
						self.client_message(locale_strings['tooManyArgs'])
					elif len(studentname) == 1:
						self.client_message(locale_strings['lastNameFirstName'])
		else:
			if conv_state == 11:
				# Resets both conversation and officer data.
				if conv_reply.lower() == "reset":
					self.client_message("Are you sure you want to reset your registered data? 'Yes' or 'no'?")
				# Returns the list of Students and subjects.
				elif conv_reply.lower() == "list":
					boysnames = "Male:\n\n"
					girlsnames = "\nFemale:\n\n"

					lastnames = []
					firstnames = []

					for names in self.officer.students_male:
						print(names)
						boysnames = boysnames + f"{names.get('last')}, {names.get('first')}\n" 
					for names in self.officer.students_female:
						print(names)
						girlsnames = girlsnames + f"{names.get('last')}, {names.get('first')}\n" 

					listnames = "List of students: \n\n"+boysnames+girlsnames+"\n\n"
					print(listnames)

					tthsubs = "TTH:\n\n"
					mwfsubs = "MWF:\n\n"

					for tsubs in self.officer.tth_subjects:
						tthsubs += tsubs+"\n"
					for msubs in self.officer.mwf_subjects:
						mwfsubs += msubs+"\n"

					subjects = "Subjects:\n\n"+tthsubs+"\n"+mwfsubs
					
					all_lists = listnames+subjects

					self.client_message(all_lists)
					self.back_state()
				# Absent marker
				elif conv_reply.split(' ', 1)[0].lower() == "a" and len(conv_reply.split(' ', 1)) > 1:
					officer = self.officer

					inputs = conv_reply.split(' ', 1)[1]
					input_name = None
					name = None
					subj_abs = None

					if '-sub' in inputs:
						subj_abs = inputs.split('-sub')[1].strip().lower()
						input_name = inputs.split('-sub')[0].strip().lower()
						name = input_name.split(',', 1)
					else:
						name = inputs.split(',', 1)

					lastname = None
					firstname = None

					print(len(name))

					if len(name) == 1:
						lastname = name[0].strip().lower()
						print(lastname)
					elif len(name) == 2:
						lastname = name[0].strip().lower()
						firstname = name[1].strip().lower()					
						print(f"{lastname}, {firstname}")

					lastnames = []
					firstnames = []

					for names in officer.students_male:
						print(names)
						lastnames.append(names.get('last').lower())
						firstnames.append(names.get('first').lower())
					for names in officer.students_female:
						print(names)
						lastnames.append(names.get('last').lower())
						firstnames.append(names.get('first').lower())

					print(lastnames)
					print(firstnames)

					if lastname in lastnames:
						lastname_count = lastnames.count(lastname)

						if lastname_count == 1:
							if subj_abs:
								if all(y in [x.lower() for x in officer.mwf_subjects] for y in [hhh.lower().strip() for hhh in subj_abs.split(',')]) or all(y in [x.lower() for x in officer.tth_subjects] for y in [hhh.lower().strip() for hhh in subj_abs.split(',')]):
									self.officer.attendance_data.mark_absent(f'{lastname}', subj_abs)
									self.client_message(f"Marked \"{name}\" absent at {subj_abs} subject.")
								else:
									self.client_message(f"Nope, the subject/s \"{subj_abs}\" does not exist in your database.\n\nType 'List' to see all registered students and subjects.")
							else:
								self.officer.attendance_data.mark_absent(f'{lastname}')
								self.client_message(f"Marked \"{name}\" absent")
						elif lastname_count > 1:
							if firstname == None:
								self.client_message(f"There are {lastname_count} students with the last name, \"{lastname}\". Please specify it with its first name. \n\nExample: \"a Llanos, Sean")
							else:
								if firstname in firstnames:
									if subj_abs:
										if all(y in [x.lower() for x in officer.mwf_subjects] for y in [hhh.lower().strip() for hhh in subj_abs.split(',')]) or all(y in [x.lower() for x in officer.tth_subjects] for y in [hhh.lower().strip() for hhh in subj_abs.split(',')]):
											self.officer.attendance_data.mark_absent(f'{lastname}', subj_abs)
											self.client_message(f"Marked \"{name}\" absent at {subj_abs} subject.")
										else:
											self.client_message(f"Nope, the subject/s \"{subj_abs}\" does not exist in your database.\n\nType 'List' to see all registered students and subjects.")
									else:
										self.officer.attendance_data.mark_absent(f'{lastname}, {firstname}')
										self.client_message(f"Marked \"{name}\" absent")
								else:
									self.client_message(f"Nope, \"{name}\" is not registered in the database.\n\nType 'List' to see all registered students and subjects.")
					else:
						self.client_message(f"\"{lastname}\" is not registered in the database.\n\nType 'List' to see all registered students and subjects.")

					self.back_state()
				# Present marker
				elif conv_reply.split(' ', 1)[0].lower() == "p" and len(conv_reply.split(' ', 1)) > 1:
					officer = self.officer

					inputs = conv_reply.split(' ', 1)[1]
					name = inputs.split(',', 1)

					lastname = None
					firstname = None

					print(len(name))

					if len(name) == 1:
						lastname = name[0].strip().lower()
						print(lastname)
					elif len(name) == 2:
						lastname = name[0].strip().lower()
						firstname = name[1].strip().lower()					
						print(f"{lastname}, {firstname}")

					lastnames = []
					firstnames = []

					for names in officer.students_male:
						print(names)
						lastnames.append(names.get('last').lower())
						firstnames.append(names.get('first').lower())
					for names in officer.students_female:
						print(names)
						lastnames.append(names.get('last').lower())
						firstnames.append(names.get('first').lower())

					print(lastnames)
					print(firstnames)

					if lastname in lastnames:
						lastname_count = lastnames.count(lastname)
						if lastname_count == 1:
							self.officer.attendance_data.mark_present(f'{lastname}')
							self.client_message(f"Marked \"{name}\" present")
						elif lastname_count > 1:
							if firstname == None:
								self.client_message(f"There are {lastname_count} students with the last name, \"{lastname}\". Please specify it with its first name. \n\nExample: \"a Llanos, Sean CD")
							else:
								if firstname in firstnames:
									self.officer.attendance_data.mark_present(f'{lastname}, {firstname}')
									self.client_message(f"Marked \"{name}\" present.")
								else:
									self.client_message(f"Nope, \"{name}\" is not registered in the database.\n\nType 'List' to see all registered students and subjects.")
					else:
						self.client_message(f"\"{lastname}\" is not registered in the database.\n\nType 'List' to see all registered students and subjects.")

					self.back_state()
				# Late marker
				elif conv_reply.split(' ', 1)[0].lower() == "l" and len(conv_reply.split(' ', 1)) > 1:
					officer = self.officer

					inputs = conv_reply.split(' ', 1)[1]
					input_name = None
					name = None
					subj_abs = None

					if '-sub' in inputs:
						subj_abs = inputs.split('-sub')[1].strip().lower()
						input_name = inputs.split('-sub')[0].strip().lower()
						name = input_name.split(',', 1)
					else:
						name = inputs.split(',', 1)

					lastname = None
					firstname = None

					print(len(name))

					if len(name) == 1:
						lastname = name[0].strip().lower()
						print(lastname)
					elif len(name) == 2:
						lastname = name[0].strip().lower()
						firstname = name[1].strip().lower()					
						print(f"{lastname}, {firstname}")

					lastnames = []
					firstnames = []

					for names in officer.students_male:
						print(names)
						lastnames.append(names.get('last').lower())
						firstnames.append(names.get('first').lower())
					for names in officer.students_female:
						print(names)
						lastnames.append(names.get('last').lower())
						firstnames.append(names.get('first').lower())

					print(lastnames)
					print(firstnames)

					if lastname in lastnames:
						lastname_count = lastnames.count(lastname)

						if lastname_count == 1:
							if subj_abs:
								if all(y in [x.lower() for x in officer.mwf_subjects] for y in [hhh.lower().strip() for hhh in subj_abs.split(',')]) or all(y in [x.lower() for x in officer.tth_subjects] for y in [hhh.lower().strip() for hhh in subj_abs.split(',')]):
									self.officer.attendance_data.mark_late(f'{lastname}', subj_abs)
									self.client_message(f"Marked \"{name}\" late at {subj_abs} subject.")
								else:
									self.client_message(f"Nope, the subject/s \"{subj_abs}\" does not exist in your database.")
							else:
								self.officer.attendance_data.mark_late(f'{lastname}')
								self.client_message(f"Marked \"{name}\" late.")
						elif lastname_count > 1:
							if firstname == None:
								self.client_message(f"There are {lastname_count} students with the last name, \"{lastname}\". Please specify it with its first name. \n\nExample: \"a Llanos, Sean")
							else:
								if firstname in firstnames:
									if subj_abs:
										if all(y in [x.lower() for x in officer.mwf_subjects] for y in [hhh.lower().strip() for hhh in subj_abs.split(',')]) or all(y in [x.lower() for x in officer.tth_subjects] for y in [hhh.lower().strip() for hhh in subj_abs.split(',')]):
											self.officer.attendance_data.mark_late(f'{lastname}', subj_abs)
											self.client_message(f"Marked \"{name}\" late at {subj_abs} subject.")
										else:
											self.client_message(f"Nope, the subject/s \"{subj_abs}\" does not exist in your database.")
									else:
										self.officer.attendance_data.mark_late(f'{lastname}, {firstname}')
										self.client_message(f"Marked \"{name}\" absent")
								else:
									self.client_message(f"Nope, \"{name}\" is not registered in the database.")
					else:
						self.client_message(f"\"{lastname}\" is not registered in the database.")

					self.back_state()
				# Returns the attendance data to the officer.
				elif conv_reply.lower() == 'attendance':
					self.client_message("Sending xlsx attendance.")
					self.officer.send_data_to_user(self.client)
					self.back_state()
				# Sends --help for commands.
				else:
					self.client_message(f"Hello {self.officer.name}.")
					self.client_message("These are the commands you can call:\n\n'list' - lists all the registered students.\n\n'l <last,first name>' - mark a student late.\n\n'a <last,first name>' - mark a student absent.\n\nNOTE: You can add in '-sub' after the students names in absent or late marker to specify the subject the student is absent or late in.\n\n'p <last,first name>' - mark a student present.\n\n'attendance' - send the attendance sheet of the current week. This is sent by default every Friday at 6PM.\n\nreset' - deletes all register and officer data saved and start again.")
					self.back_state()
			# Confirmation for reset.
			if conv_state == 12:
				if conv_reply.lower() == "yes":
					self.client_message("Resetting...")

					json_data = self.officer.recipient_id

					os.remove(get_dir(f"data/officer_data/officer-{json_data}.json"))

					if json_data in register_load_data:
						register_load_data.pop(f"{json_data}")

					file_data = {}

					with open(get_dir('data/register.json'), 'r') as register_data:
						file_data = json.load(register_data)

					if len(file_data) <= 1:
						os.remove(get_dir('data/register.json'))
					else:
						with open(get_dir('data/register.json'), 'w') as save_data:
							file_data.pop(f"{json_data}")
							save_data.seek(0)
							json.dump(file_data, save_data, indent=4)
							save_data.truncate()
							print("Register data was successfully loaded, modified and saved.")

					self.client.conversations.discard(self.client.is_in_conversations(self.recipient_id).get('conv'))
					self.client_message("Successfully reset. Type anything to continue and register.")
					del self
				else:
					self.client_message("Oh okay then.")
					self.back_state(10)
			# Daily notification for attendance
			elif conv_state == 100:
				self.client_message(f"Hello {self.officer.name}, can you kindly check today's attendance please?")
				self.back_state(10)

	def reply(self, message_string):
		if not self.is_typing:
			self.conv_state += 1
			self.conv_reply = message_string
			self.convo()

	def client_message(self, message_string):
		self.client.send_text_message(self.recipient_id, message_string)

	def back_state(self, state=None):
		if state == None:
			self.conv_state -= 1
		else:
			if RepresentsInt(state):
				self.conv_state = state
			else:
				print("Conv_state is an int! Smh")

# Subclass of fbchat.Client and override required methods
class AmsBot(Bot):
	def set_locale_strings(self, lang):
		if lang == 'en':
			with open(get_dir('locale/en.json')) as f:
				self.locale_strings = json.load(f)
		elif lang == 'bs':
			with open(get_dir('locale/bisaya.json')) as f:
				self.locale_strings = json.load(f)

	def get_locale_strings(self, lang):
		locale_strings = {}
		if lang == 'en':
			with open(get_dir('locale/en.json')) as f:
				locale_strings = json.load(f)
		elif lang == 'bs':
			with open(get_dir('locale/bisaya.json')) as f:
				locale_strings = json.load(f)
		return locale_strings

	def initialize(self):
		self.set_locale_strings('en')
		self.conversations = set()

	def on_message(self, recipient_id=None, message_string=None):
		locale_strings = self.locale_strings

		if message_string == ".test":
			self.send_text_message(recipient_id, locale_strings['greetings'])
			print("Welcomed @"+recipient_id+" at "+f"{time.strftime('%X')}")
		elif message_string == ".testattendance":
			print(recipient_id)
			self.send_file(recipient_id=recipient_id, file_path=get_dir("data/attendance.xlsx"))
			print("Sent attendance file "+f"{time.strftime('%X')}")
		else:
			self.interact_convo(recipient_id, message_string)
		print("Interacting with a user.")
		print(f"No. of conversations: {len(self.conversations)}")

	# This function allows the bot to interact with the current author
	def interact_convo(self, recipient_id, message_string):
		if self.is_in_conversations(recipient_id).get("bool"):
			conv = self.is_in_conversations(recipient_id).get("conv")
			conv.reply(message_string)
			print("Replied "+recipient_id)
		elif recipient_id in register_load_data:
			loaded_officer = Officer(recipient_id)
			loaded_officer.load()
			sched.add_job(func=loaded_officer.send_data_to_user, trigger='cron', args=[self], day_of_week='mon', hour=18, minute=00)
			c = Conversation(self, None, loaded_officer)
			c.reply(message_string)
			self.conversations.add(c)
			print("Loaded officer data and has been conversed with.")
		else:
			c = Conversation(self, recipient_id)
			c.convo()
			self.conversations.add(c)
			print("Started conversation at "+f"{time.strftime('%X')}")

	# This function checks if the interacting author is having a conversation in the current session and
	# also returns the Conversation object.
	def is_in_conversations(self, recipient_id):
		in_conversations = False
		conv = None

		for conversation in self.conversations:
			if recipient_id == conversation.recipient_id:
				in_conversations = True
				conv = conversation
				break

		this = {"bool" : in_conversations, "conv" : conv}

		return this

# Initialization of the app, listener and send API
app = Flask(__name__)

with open(get_dir("config/config.json")) as auth_file:
	auth = json.load(auth_file)

ACCESS_TOKEN = auth['ACCESS_TOKEN']
VERIFY_TOKEN = auth['VERIFY_TOKEN']

bot = AmsBot(ACCESS_TOKEN)
bot.initialize()

# Scheduled jobs of the bot
sched = BackgroundScheduler()

register_load_data = {}

def schedule_officer(recipient_id):
	c = None
	if bot.is_in_conversations(recipient_id).get('bool'):
		c = bot.is_in_conversations(recipient_id).get('conv')
	else:
		loaded_officer = Officer(recipient_id)
		loaded_officer.load()
		c = Conversation(bot, None, loaded_officer)
		bot.conversations.add(c)
	c.conv_state = 100
	c.convo()
	print(f"Reminding: {recipient_id} for the daily attendance.")

if os.path.exists(get_dir('data/register.json')):
	with open(get_dir('data/register.json')) as loaded_register_data:
		register_load_data = json.load(loaded_register_data)
	for registered_schedule in register_load_data:
		# MWF
		print(f"recipient_id load: {registered_schedule}")
		time_mwf = register_load_data[registered_schedule].get('mwf_time').split(':')
		hour_mwf = int(time_mwf[0])
		minute_mwf = int(time_mwf[1])
		print(time_mwf)
		print(hour_mwf)
		print(minute_mwf)

		sched.add_job(func=schedule_officer, trigger='cron', args=[registered_schedule], day_of_week='mon', hour=hour_mwf, minute=minute_mwf)
		sched.add_job(func=schedule_officer, trigger='cron', args=[registered_schedule], day_of_week='wed', hour=hour_mwf, minute=minute_mwf)
		sched.add_job(func=schedule_officer, trigger='cron', args=[registered_schedule], day_of_week='fri', hour=hour_mwf, minute=minute_mwf)

		# TTH
		time_tth = register_load_data[registered_schedule].get('tth_time').split(':')
		hour_tth = int(time_tth[0])
		minute_tth = int(time_tth[1])
		print(time_tth)
		print(hour_tth)
		print(minute_tth)

		sched.add_job(func=schedule_officer, trigger='cron', args=[registered_schedule], day_of_week='tue', hour=hour_tth, minute=minute_tth)
		sched.add_job(func=schedule_officer, trigger='cron', args=[registered_schedule], day_of_week='thu', hour=hour_tth, minute=minute_tth)

sched.start()

atexit.register(lambda: sched.shutdown(wait=False))

# Listener base code of the bot
@app.route("/getpost", methods=['GET', 'POST'])
def receive_message():
	if request.method == 'GET':
		""" Before allowing people to message your bot, Facebook has implemented a verify token
		that confirms all requests that your bot receives came from Facebook. """
		token_sent = request.args.get("hub.verify_token")
		return verify_fb_token(token_sent)
	# if the request was not get, it must be POST and we can just proceed with sending a message back to user
	elif request.method == 'POST':
		#  get whatever message a user sent the bot
		output = request.get_json()
		print(f"{output}")
		for event in output['entry']:
			messaging = event['messaging']
			for message in messaging:
				recip = message['sender']['id']
				# if user sends a text
				if message['message'].get('text'):
					bot.on_message(recip, message_string=message['message']['text'])
				# if user sends us a GIF, photo,video, or any other non-text item
				if message['message'].get('attachments'):
					bot.on_message(recip, get_message())
	return "Message Processed"

def verify_fb_token(token_sent):
	# take token sent by facebook and verify it matches the verify token you sent
	# if they match, allow the request, else return an error
	if token_sent == VERIFY_TOKEN:
		return request.args.get("hub.challenge")
	return 'Invalid verification token'

# This is the page for the web interface
@app.route("/")
def status():
	bot_officers = list()	
	bot_conv_reps = list()
	bot_conv_states =  list()

	bot_convs = ""

	for convo in bot.conversations:
		bot_officers.append(convo.officer.name)
		bot_conv_reps.append(convo.conv_reply)
		bot_conv_states.append(convo.conv_state)

	for x in range(len(bot.conversations)):
		bot_convs += f"[|{bot_officers[x]}|	Reply: \"{bot_conv_reps[x]}\"	State: {bot_conv_states[x]}]"

	bot_status = f"""<h1>No. of conversations: {len(bot.conversations)}
	Conversations: 

	{bot_convs}<h1>
	"""

	if bot_status:
		return bot_status
	else:
		return "No status to display yet."

# Get user profile info
def get_profile(PSID):
	r = requests.get(f"https://graph.facebook.com/{PSID}?fields=first_name,last_name,profile_pic,gender&access_token={ACCESS_TOKEN}")
	return r.json()

# chooses a random message to send to the user
def get_message():
	sample_responses =["Char ahahahaha", "Hahahaha sanaol"]
	#  return selected item to the user
	return random.choice(sample_responses)

if __name__ == "__main__":
	app.run()
